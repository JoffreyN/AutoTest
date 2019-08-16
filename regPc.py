# from _mails import *
from tools import *
try:
	from menhu.config import *
except ModuleNotFoundError:
	from config import *
# from PIL import Image
from uuid import uuid1,uuid4
# from OMlogin import loginOM
from selenium import webdriver
from random import randint,choice
from RecognizeCode import GetCode
from precredit import uploadExcel,addWhiteList
from Login import login,changePassword
from SQLreg import mysqlOpt
from checkBboss import bbossCheck
from AM import AutoFill
from H5Login import loginH5,changePasswordH5
import sys,time,os,requests,io,re,openpyxl,argparse
from OMquerier import getOMpassword,getMerchantNum
from genRegData import encryptRegData
from addOperator import addOperator
from _AMelect import AutoFillElect
from _area import getArea
from _FactoringSCF import *
from MMcredit import addUser,entryCoreCompany
from isBuilding import checkBuild
# fake=faker.Faker('zh_CN')
# user_mail,password=choice(accountPool)

# #########################发送注册邮件#######start#####################################
# def SendEmail():#发送注册邮件
# 	head=GetHead()
# 	while True:
# 		code=GetRegisterCode(head)
# 		if ChickCode(code,head)=='"1"':break
# 	if saveToMysql:mysqlOpt(f"""INSERT INTO registerinfo (CreateDate,Environmental,loginID) VALUES ("{time.strftime('%Y-%m-%d %X',time.localtime())}","PC {envDict[args.env]}","{company.loginID}")""")
# 	else:save_file(f'登录号： {company.loginID}')
# 	print(f'登录号： {company.loginID}')
# 	r=requests.post(f'{domain}/bestpay/validateLogInID',headers=head,data=f'entLoginName={company.loginID}')
# 	key=f'cashier=&regUserName=&logInID={company.loginID}&emailName={user_mail}&securityCode_ui={code}&securityCode={code}&isEmailRegistered=false'
# 	sys.stdout.write('正在发送注册邮件...\t\t\t\t\r')
# 	sys.stdout.flush()
# 	r=requests.post(f'{domain}/bestpay/doRegister',headers=head,data=key)
# 	if '"success":true' in r.text:return True
# 	else:
# 		print(r.text)
# 		sys.exit(0)
# 		return False

# def ChickCode(code,head):#检查验证码是否正确
# 	r=requests.post(f'{domain}/securityCode/validateCode?securityCode={code}',headers=head)
# 	return r.text

# def GetHead():#获取cookie，请求头
# 	head={'User-Agent':'Mozilla/5.0 (Windows NT 10.0; WOW64; rv:49.0) Gecko/20100101 Firefox/49.0','Content-Type':'application/x-www-form-urlencoded; charset=UTF-8'}
# 	r=requests.post(f'{domain}/register/auth',headers=head,data=f'id=&hasPlatCode=Y&platCode={platCode}')
# 	if r.text=='{"result":true,"errorCode":"","errorMsg":"","success":true}':
# 		head['cookie']=f"{[i for i in r.headers['set-cookie'].split(';') if i.lower().startswith('session')][0]};platCode={platCode};"
# 		return head
# 	else:print(r.text);sys.exit(0)

# def GetRegisterCode(head):
# 	sys.stdout.write('正在识别验证码...\t\t\t\t\r')
# 	sys.stdout.flush()
# 	img=io.BytesIO(requests.get(f'{domain}/securityCode/imageCaptcha',headers=head).content)
# 	code=GetCode.GetCode_register(img)
# 	return code	
# #############################发送注册邮件### End ################################################################

# def GetRegisterURL(pop3_server='pop.qq.com'):#从邮件中读取注册链接
# 	num0=countMails(pop3_server,user_mail,password)
# 	if SendEmail():
# 		sys.stdout.write(f'正在提取{user_mail}邮件中注册链接...\r')		
# 		sys.stdout.flush()
# 		count=0
# 		while True:
# 			count+=1
# 			num1=countMails(pop3_server,user_mail,password)
# 			if num1==num0+1:						
# 				texts=getMailText(Get_msg(pop3_server,user_mail,password))
# 				try:					
# 					RegisterURL=[i for i in re.findall(r"<a[^<>]*?>\s*(.*?)\s*</a>",texts) if i.startswith('http')][0]
# 				except Exception as e:
# 					print(e,texts)
# 				if args.env=='p':
# 					RegisterURL=RegisterURL.replace('https://f','https://f-pre')
# 				elif args.env=='k':
# 					RegisterURL=RegisterURL.replace('http://116.228.151.161:18178','http://f.test1.bestpay.net')
# 				return RegisterURL
# 			if count==12:
# 				print('邮件链接获取失败！\t\t\t\t\t')
# 				myexit()
# 			time.sleep(5)

################开始填表#############################################
def openurl(url,**bro):
	# sys.stdout.write('即将开始填表......\t\t\t\t\r')
	# sys.stdout.flush()
	chrome_option=webdriver.ChromeOptions()
	chrome_option.add_argument('disable-infobars')#禁止显示“Chrome正受到自动化软件的控制”
	# chrome_option.add_argument("--remote-debugging-port=0")
	if noPic:
		chrome_option.add_experimental_option("prefs", {"profile.managed_default_content_settings.images": 2})#谷歌无图模式
		# chrome_option.add_argument('-headless') #谷歌无头模式
	browser=webdriver.Chrome(options=chrome_option) if not bro else bro['bro']
	# browser.implicitly_wait(10)
	browser.get(url)
	browser.maximize_window()
	return browser

################第一步#############################第一步#############################第一步#############################第一步#####################
#企业商户（包含个体工商户）
def Putong(browser):#选择普通营业执照
	time.sleep(3)
	if args.robot:saveFileRobot('开始填写开户资料')
	browser=waitTo(browser,SCENSEE,way='id',name='certificateType1',operate='click',timeout=30)#选择普通营业执照
	if args.overfill:overFill(browser,'enterpriseName',company.companyName,comId='normalCertificateNo')
	else:browser.find_element_by_id('enterpriseName').send_keys(company.companyName)
	if args.overfill:overFill(browser,'normalAccName1',company.companyName,comId='normalCertificateNo')
	else:browser.find_element_by_id('normalAccName1').send_keys(company.companyName)#经营名称/简称
	#随机选择行业
	browser.find_element_by_id('trdTpCdType1_combobox_default').click()
	if type(1)==type(company.industry):
		browser.find_element_by_xpath(f'//*[@id="trdTpCdType1_combobox_ul"]/li[{company.industry}]').click()
	else:
		company.industry=company.industry if company.industry in industryList else '其他'
		browser.find_element_by_xpath(f'//*[@id="trdTpCdType1_combobox_ul"]/li[text()="{company.industry}"]').click()
	#随机选择区域
	browser.find_element_by_id('province1_combobox_default').click()
	if type(1)==type(company.province):
		browser.find_element_by_xpath(f'//*[@id="province1_combobox_ul"]/li[{company.province}]').click()
	else:
		browser.find_element_by_xpath(f'//*[@id="province1_combobox_ul"]/li[text()="{company.province}"]').click()#点击省
		browser.find_element_by_id('city2_combobox_default').click()
		browser.find_element_by_xpath(f'//*[@id="city1_combobox_ul"]/li[text()="{company.city}"]').click()#点击市
		browser.find_element_by_id('areaCode1_combobox_default').click()
		browser.find_element_by_xpath(f'//*[@id="areaCode1_combobox_ul"]/li[text()="{company.district}"]').click()#点击区
	if args.overfill:overFill(browser,'normalRunAddr',company.officeAddr,comId='normalCertificateNo')
	else:browser.find_element_by_id('normalRunAddr').send_keys(company.officeAddr)#办公地址
	if saveToMysql:mysqlOpt(f"UPDATE registerinfo SET RegistNum='{company.BLRN}' WHERE LoginID='{company.loginID}'")
	else:save_file(f'营业执照注册号： {company.BLRN}')
	print(f'企业名称：{company.companyName}\n营业执照注册号： {company.BLRN}')
	if args.robot:saveFileRobot(f'企业名称：{company.companyName}\n营业执照注册号： {company.BLRN}')
	browser.find_element_by_id('normalCertificateNo').send_keys(company.BLRN)#营业执照注册号,15位数字,随机生成
	# dates=company.regTime
	browser.find_element_by_id('licenseBeginDate').send_keys(company.regTime)#营业执照有效期，开始
	if re.match(r'\d{4}-\d{2}-\d{2}',company.endTime):
		browser.find_element_by_id('licenseEndDate').send_keys(company.endTime)#营业执照有效期，结束
	else:
		browser.find_element_by_id('bizLicenseValidTime1').click()#营业执照有效期，长期
	browser.find_element_by_id('orgCode').send_keys(company.orgCode)#组织机构代码
	browser.find_element_by_id('orgCodeBeginDate').send_keys(company.regTime)#组织机构代码有效期，开始
	if re.match(r'\d{4}-\d{2}-\d{2}',company.endTime):
		browser.find_element_by_id('orgCodeEndDate').send_keys(company.endTime)#组织机构代码有效期，结束
	else:
		browser.find_element_by_id('orgCodeValidTime1').click()#组织机构代码有效期，长期
	browser.find_element_by_id('taxRegCode').send_keys(company.taxRegCode)#税务登记证号,15位数字或字母，随机生成
	browser.find_element_by_id('bankApprovedNum').send_keys(company.permitNum)#开户许可证核准号，14位数字或字母，随机生成
	if args.overfill:overFill(browser,'normalWebInfo',company.fake.url(),comId='normalCertificateNo')
	else:browser.find_element_by_id('normalWebInfo').send_keys(company.fake.url())#网络支付商户网址信息
	if args.overfill:overFill(browser,'normalOperate',company.scope,comId='normalCertificateNo')
	else:browser.find_element_by_id('normalOperate').send_keys(company.scope)#经营范围
	# if args.overfill:overFill(browser,'normalRegAmt',str(company.regCapital),comId='normalCertificateNo')
	# else:
	browser.find_element_by_id('normalRegAmt').send_keys(company.regCapital)#注册资本金	
	return browser
def Sanzhengheyi(browser):#选择三证合一营业执照
	time.sleep(3)
	if args.robot:saveFileRobot('开始填写开户资料')
	browser=waitTo(browser,SCENSEE,way='id',name='certificateType2',operate='click',timeout=30)#选择三证合一营业执照
	if args.overfill:overFill(browser,'usccCustomerName',company.companyName,comId='usccCodeCertificateNo')
	else:browser.find_element_by_id('usccCustomerName').send_keys(company.companyName)#企业名称
	if args.overfill:overFill(browser,'usccAccName1',company.companyName,comId='usccCodeCertificateNo')
	else:browser.find_element_by_id('usccAccName1').send_keys(company.companyName)#经营名称/简称
	if saveToMysql:mysqlOpt(f"UPDATE registerinfo SET USCNUM='{company.BLRN}' WHERE LoginID='{company.loginID}'")
	else:save_file(f'统一社会信用代码： {company.BLRN}')
	print(f'企业名称：{company.companyName}\n统一社会信用代码： {company.BLRN}')
	if args.robot:saveFileRobot(f'企业名称：{company.companyName}\n统一社会信用代码： {company.BLRN}')
	browser.find_element_by_id('usccCodeCertificateNo').send_keys(company.BLRN)#统一社会信用代码,18位数字或字母，随机生成
	#随机选择区域
	browser.find_element_by_id('province2_combobox_default').click()
	if type(1)==type(company.province):
		browser.find_element_by_xpath(f'//*[@id="province2_combobox_ul"]/li[{company.province}]').click()
	else:
		browser.find_element_by_xpath(f'//*[@id="province2_combobox_ul"]/li[text()="{company.province}"]').click()#点击省
		browser.find_element_by_id('city2_combobox_default').click()
		browser.find_element_by_xpath(f'//*[@id="city2_combobox_ul"]/li[text()="{company.city}"]').click()#点击市
		browser.find_element_by_id('areaCode2_combobox_default').click()
		browser.find_element_by_xpath(f'//*[@id="areaCode2_combobox_ul"]/li[text()="{company.district}"]').click()#点击区
	if args.overfill:overFill(browser,'usccRunAddr',company.officeAddr,comId='usccCodeCertificateNo')
	else:browser.find_element_by_id('usccRunAddr').send_keys(company.officeAddr)#办公地址
	#随机选择行业
	browser.find_element_by_id('trdTpCdType2_combobox_default').click()
	if type(1)==type(company.industry):
		browser.find_element_by_xpath(f'//*[@id="trdTpCdType2_combobox_ul"]/li[{company.industry}]').click()
	else:
		company.industry=company.industry if company.industry in industryList else '其他'
		browser.find_element_by_xpath(f'//*[@id="trdTpCdType2_combobox_ul"]/li[text()="{company.industry}"]').click()
	browser.find_element_by_id('usccLicenseBeginDate').send_keys(company.regTime)#营业执照有效期，开始
	if re.match(r'\d{4}-\d{2}-\d{2}',company.endTime):
		browser.find_element_by_id('usccLicenseEndDate').send_keys(company.endTime)#营业执照有效期，结束
	else:
		browser.find_element_by_id('bizLicenseValidTime2').click()#勾选长期
	# if saveToMysql:mysqlOpt(f"UPDATE registerinfo SET PermitNum='{ACCN}' WHERE LoginID='{company.loginID}'")
	# else:save_file(f'开户许可证核准号： {ACCN}')
	browser.find_element_by_id('usccBankApproved').send_keys(company.permitNum)#开户许可证核准号，14位数字或字母，随机生成
	if args.overfill:overFill(browser,'usccWebInfo',company.fake.url(),comId='usccCodeCertificateNo')
	else:browser.find_element_by_id('usccWebInfo').send_keys(company.fake.url())#网络支付商户网址信息
	if args.overfill:overFill(browser,'usccOperate',company.scope,comId='usccCodeCertificateNo')
	else:browser.find_element_by_id('usccOperate').send_keys(company.scope)#经营范围
	# if args.overfill:overFill(browser,'usccRegAmt',str(company.regCapital),comId='normalCertificateNo')
	# else:
	browser.find_element_by_id('usccRegAmt').send_keys(company.regCapital)#注册资本金
	return browser
def Getigongshanghu(browser):#选择个体工商户营业执照
	time.sleep(3)
	if args.robot:saveFileRobot('开始填写开户资料')
	browser=waitTo(browser,SCENSEE,way='id',name='certificateType3',operate='click',timeout=30)#选择个体工商户营业执照
	if args.overfill:overFill(browser,'singleCustomerName',company.companyName,comId='singleCertificateNo')
	else:browser.find_element_by_id('singleCustomerName').send_keys(company.companyName)#商户名称
	if args.overfill:overFill(browser,'singleAccName1',company.companyName,comId='singleCertificateNo')
	else:browser.find_element_by_id('singleAccName1').send_keys(company.companyName)#经营名称/简称	
	if saveToMysql:mysqlOpt(f"UPDATE registerinfo SET RegistNum='{company.BLRN}' WHERE LoginID='{company.loginID}'")
	else:save_file(f'营业执照注册号： {company.BLRN}')
	print(f'企业名称：{company.companyName}\n营业执照注册号： {company.BLRN}')
	if args.robot:saveFileRobot(f'企业名称：{company.companyName}\n营业执照注册号： {company.BLRN}')
	browser.find_element_by_id('singleCertificateNo').send_keys(company.BLRN)#营业执照注册号,15至20个字符,随机生成
	#随机选择区域
	browser.find_element_by_id('province3_combobox_default').click()
	if type(1)==type(company.province):
		browser.find_element_by_xpath(f'//*[@id="province3_combobox_ul"]/li[{company.province}]').click()
	else:
		browser.find_element_by_xpath(f'//*[@id="province3_combobox_ul"]/li[text()="{company.province}"]').click()#点击省
		browser.find_element_by_id('city3_combobox_default').click()
		browser.find_element_by_xpath(f'//*[@id="city3_combobox_ul"]/li[text()="{company.city}"]').click()#点击市
		browser.find_element_by_id('areaCode3_combobox_default').click()
		browser.find_element_by_xpath(f'//*[@id="areaCode3_combobox_ul"]/li[text()="{company.district}"]').click()#点击区	
	if args.overfill:overFill(browser,'singleRunAddr',company.officeAddr,comId='singleCertificateNo')
	else:browser.find_element_by_id('singleRunAddr').send_keys(company.officeAddr)#办公地址
	#随机选择行业
	browser.find_element_by_id('trdTpCd_select3').click()
	if type(1)==type(company.industry):
		browser.find_element_by_xpath(f'//*[@id="trdTpCdType3_combobox_ul"]/li[{company.industry}]').click()
	else:
		company.industry=company.industry if company.industry in industryList else '其他'
		browser.find_element_by_xpath(f'//*[@id="trdTpCdType3_combobox_ul"]/li[text()="{company.industry}"]').click()
	browser.find_element_by_id('singleLicenseBeginDate').send_keys(company.regTime)#营业执照有效期，开始
	if re.match(r'\d{4}-\d{2}-\d{2}',company.endTime):
		browser.find_element_by_id('singleLicenseEndDate').send_keys(company.endTime)#营业执照有效期，结束
	else:
		browser.find_element_by_id('bizLicenseValidTime3').click()#勾选长期
	if args.overfill:overFill(browser,'singleOperate',company.scope,comId='singleCertificateNo')
	else:browser.find_element_by_id('singleOperate').send_keys(company.scope)#经营范围
	# browser.find_element_by_id('usccRegAmt').send_keys(company.regCapital)#注册资本金	
	return browser

def legalinformation(browser):#法人信息
	time.sleep(3)
	if args.robot:saveFileRobot('开始填写法人信息')
	browser.find_element_by_id('legalName').send_keys(company.legalName)#法人代表姓名
	#选择回乡证
	#browser.find_element_by_id('legalCertType_combobox_default').click()
	#browser.find_elements_by_class_name('ui-select-item')[169].click()
	browser.find_element_by_id('paper11').send_keys(company.legalID)#证件号码
	browser.find_element_by_id('legalIdCardBeginDate').send_keys(company.LICST)#证件有效期，开始
	if re.match(r'\d{4}-\d{2}-\d{2}',company.LICET):
		browser.find_element_by_id('legalIdCardEndDate').send_keys(company.LICET)#证件有效期，结束
	else:
		browser.find_element_by_id('idCardValidTime').click()#勾选长期
	browser.find_element_by_id('legalPhone').send_keys(company.legalTel)#手机号码
	print(f'法人手机号： {company.legalTel}\t\t\t\t')
	#browser.find_element_by_id('proxy2').click()#我是代理人
	# os.system('pause')
	browser.find_element_by_xpath('//*[@id="main"]/div[3]/div/input[2]').click()#下一步
	time.sleep(2)
	return browser

def beneficiaryInfo(browser):
	if args.robot:saveFileRobot('开始填写受益人信息')
	# if args.env=='s':return browser
	if args.type in ['1','2']:
		if args.mode=='c':
			time.sleep(2)
			browser.find_element_by_name('beneficiaryName1').send_keys(company.beneName)#受益人名称
		else:
			browser=waitTo(browser,SCENSEE,way='name',name='beneficiaryName1',operate='send_keys',value=company.beneName)#受益人名称
		browser.find_element_by_id('beneficiaryIdNo1').send_keys(company.beneID)#证件号码
		browser.find_element_by_id('idEffectiveDateBegin1').send_keys(company.BICST)#证件有效期，开始
		if re.match(r'\d{4}-\d{2}-\d{2}',company.BICET):
			browser.find_element_by_id('idEffectiveDate1').send_keys(company.BICET)#证件有效期，结束
		else:
			browser.find_element_by_id('bizLicenseValidTime1').click()#证件有效期，勾选长期
	if args.overfill:overFill(browser,'eneficiaryAddress1',company.beneAddr,comId='beneficiaryIdNo1')
	else:browser=waitTo(browser,SCENSEE,way='id',name='eneficiaryAddress1',operate='send_keys',value=company.beneAddr)#受益人地址
	# myexit()
	browser.find_element_by_css_selector('#main>div:nth-child(3)>div>input:nth-child(2)').click()#下一步
	time.sleep(2)
	return browser

def Geti(browser):#选择个人商户
	browser.find_element_by_id('registerType_combobox_default').click()
	browser.find_elements_by_class_name('ui-select-item')[1].click()
	browser.find_element_by_id('individualName').send_keys(company.legalName)#姓名
	#选择回乡证
	#browser.find_element_by_id('registerCertType_combobox_default').click()
	#browser.find_element_by_xpath('//*[@id="registerCertType_combobox_ul"]/li[1]').click()
	browser.find_element_by_id('individualIdCard').send_keys(company.legalID)#证件号码
	browser.find_element_by_id('individualIdCardBeginDate').send_keys(company.regTime)#证件有效期，开始
	browser.find_element_by_id('individualIdCardEndDate').send_keys(company.endTime)#证件有效期，结束
	#随机选择区域
	browser.find_element_by_id('province4_combobox_default').click()
	browser.find_element_by_xpath('//*[@id="province4_combobox_ul"]/li[%d]'%randint(2,33)).click()
	browser.find_element_by_id('individualRunAddr').send_keys(company.officeAddr)#办公地址
	#随机选择所属行业
	browser.find_element_by_id('trdTpCdType4_combobox_default').click()
	browser.find_element_by_xpath('//*[@id="trdTpCdType4_combobox_ul"]/li[%d]'%randint(1,9)).click()
	#随机选择职业
	browser.find_element_by_id('individualProfession_combobox_default').click()
	browser.find_element_by_xpath('//*[@id="individualProfession_combobox_ul"]/li[%d]'%randint(1,8)).click()
	browser.find_element_by_id('individualPhone').send_keys(company.legalTel)#手机号码
	browser.find_element_by_xpath('//*[@id="main"]/div[3]/div/input[2]').click()#下一步
	time.sleep(2)
	return browser

################第二步#############################第二步#############################第二步#############################第二步#####################
def postpic0(browser):
	if args.robot:saveFileRobot('开始上传图片')
	if args.mode=='c':
		browser.find_element_by_id('bizLicensePhoto').send_keys(company.bizLicensepath)
	else:
		browser=waitTo(browser,SCENSEE,way='id',name='bizLicensePhoto',operate='send_keys',value=company.bizLicensepath)#营业执照
	time.sleep(1)
	browser.find_element_by_id('orgCodePhoto').send_keys(company.orgCodepath)#组织机构代码证
	time.sleep(1)
	browser.find_element_by_id('taxRegCodePhoto').send_keys(company.taxRegpath)#税务登记证
	time.sleep(1)
	browser.find_element_by_id('openAcctPermitPhoto').send_keys(company.openAcctpath)#银行开户许可证
	time.sleep(1)
	browser.find_element_by_id('legalIdNumTPhoto').send_keys(company.legalApath)#法人身份证正面
	time.sleep(1)
	browser.find_element_by_id('legalIdNumBPhoto').send_keys(company.legalBpath)#法人身份证反面
	time.sleep(1)
	browser.find_element_by_id('handHeldIDPhoto').send_keys(company.handIDpath)#手持身份证正面照
	if args.noCheck:
		waite(10,10,'等待上传图片')
		# time.sleep(10)
	else:time.sleep(4)
	browser.find_element_by_xpath('//*[@id="main"]/div[3]/a[2]').click()#下一步
	time.sleep(2)
	return browser

def postpic1(browser):
	if args.robot:saveFileRobot('开始上传图片')
	if args.mode=='c':
		browser.find_element_by_id('bizLicensePhoto').send_keys(company.bizLicensepath)#三证合一营业执照
	else:
		browser=waitTo(browser,SCENSEE,way='id',name='bizLicensePhoto',operate='send_keys',value=company.bizLicensepath)#营业执照
	time.sleep(1)
	browser.find_element_by_id('openAcctPermitPhoto').send_keys(company.openAcctpath)#银行开户许可证
	time.sleep(1)
	browser.find_element_by_id('legalIdNumTPhoto').send_keys(company.legalApath)#法人身份证正面
	time.sleep(1)
	browser.find_element_by_id('legalIdNumBPhoto').send_keys(company.legalBpath)#法人身份证反面
	time.sleep(1)
	browser.find_element_by_id('handHeldIDPhoto').send_keys(company.handIDpath)#手持身份证正面照
	if args.noCheck:
		waite(10,10,'等待上传图片')
		# time.sleep(10)
	else:time.sleep(4)
	browser.find_element_by_xpath('//*[@id="main"]/div[3]/a[2]').click()#下一步
	time.sleep(2)
	return browser

def postpic2(browser):
	if args.robot:saveFileRobot('开始上传图片')
	if args.mode=='c':
		browser.find_element_by_id('bizLicensePhoto').send_keys(company.bizLicensepath)#三证合一营业执照
	else:
		browser=waitTo(browser,SCENSEE,way='id',name='bizLicensePhoto',operate='send_keys',value=company.bizLicensepath)#营业执照
	time.sleep(1)
	browser.find_element_by_id('legalIdNumTPhoto').send_keys(company.legalApath)#法人身份证正面
	time.sleep(1)
	browser.find_element_by_id('legalIdNumBPhoto').send_keys(company.legalBpath)#法人身份证反面
	time.sleep(1)
	browser.find_element_by_id('handHeldIDPhoto').send_keys(company.handIDpath)#手持身份证正面照
	if args.noCheck:
		waite(10,10,'等待上传图片')
		# time.sleep(10)
	else:time.sleep(4)
	browser.find_element_by_xpath('//*[@id="main"]/div[3]/a[2]').click()#下一步
	time.sleep(2)
	return browser

def postpic3(browser):
	if args.robot:saveFileRobot('开始上传图片')
	if args.mode=='c':
		browser.find_element_by_id('legalIdNumTPhoto').send_keys(company.legalApath)#法人身份证正面
	else:
		browser=waitTo(browser,SCENSEE,way='id',name='legalIdNumTPhoto',operate='send_keys',value=company.legalApath)#营业执照
	time.sleep(1)
	browser.find_element_by_id('legalIdNumBPhoto').send_keys(company.legalBpath)#法人身份证反面
	time.sleep(1)
	browser.find_element_by_id('handHeldIDPhoto').send_keys(company.handIDpath)#手持身份证正面照
	if args.noCheck:
		waite(10,10,'等待上传图片')
		# time.sleep(10)
	else:time.sleep(4)
	browser.find_element_by_xpath('//*[@id="main"]/div[3]/a[2]').click()#下一步
	time.sleep(2)
	return browser

################第三步#############最后一步################第三步#############最后一步################第三步#############最后一步################
def fillbankinfor(browser,code):
	if args.robot:saveFileRobot('开始填写银行卡信息')
	args.skip=1
	if args.skip:
		browser.find_element_by_id('bindBankCard1').click()#跳过,下次再进行绑定银行卡
	else:
		if code in [0,1]:#前两种需要填开户户名
			browser=waitTo(browser,SCENSEE,way='id',name='bankAcctName',operate='send_keys',value=company.bankAccName)#开户户名
			# browser.find_element_by_id('bankAcctName').send_keys(company.bankAccName)
		browser=waitTo(browser,SCENSEE,way='id',name='bankAcctno',operate='send_keys',value=company.bankCode)#银行账号
		# browser.find_element_by_id('bankAcctno').send_keys(company.bankCode)
		browser.find_element_by_id('bankCode_combobox_default').click()#点击开户银行
		if type(1)!=type(company.bankName):
			try:
				time.sleep(2)
				browser.find_element_by_xpath(f'//*[@id="bankCode_combobox_ul"]/li[text()="{company.bankName}"]').click()#点击银行名称
			except SCENSEE:
				print('未找到该银行：',company.bankName)
		#随机选择开户银行所在地
		browser.find_element_by_id('bankAreaProvinceCode_combobox_default').click()
		browser.find_element_by_xpath(f'//*[@id="bankAreaProvinceCode_combobox_ul"]/li[{company.bankProvince}]').click()
		if args.overfill:overFill(browser,'bankBranchName',company.bankBranchName,comId='bankAcctno')
		else:browser.find_element_by_id('bankBranchName').send_keys(company.bankBranchName)#开户支行名称
		#开户证件类型
		browser.find_element_by_id('bankCertType_combobox_default').click()
		browser.find_element_by_xpath('//*[@id="bankCertType_combobox_ul"]/li[2]').click()#身份证
		browser.find_element_by_id('bankCertNo').send_keys(company.legalID)#证件号码
		browser.find_element_by_id('bankReservePhone').send_keys(company.legalTel)#银行开户手机号码
		browser.find_element_by_id('bankCardPhoto').send_keys(bankCardpath)#对公银行开户回执单
		time.sleep(2)
	# os.system('pause')
	browser.find_element_by_id('nextStepBtn').click()#下一步
	while True:
		try:
			if '将在2个工作日内完成审核' in browser.find_element_by_xpath('/html/body/div[3]/div[3]/div[2]/p[1]').text:
				return browser
		except SCENSEE:
			time.sleep(1)
			continue

def getParserPC():
	parser=argparse.ArgumentParser(description='程序功能：\n    1、PC端注册+修改密码+预授信+新资质审核+申请白条;\n    2、电子凭证开户+授信+新增角色、用户',formatter_class=argparse.RawTextHelpFormatter)
	parser.add_argument('-m',dest='mode',help="开户模式（默认 a）：\n    a: 普通模式\n    b: 电子凭证模式\n    c: 小CEO批量开户",required=False,default='a',choices=['a','b','c'])
	parser.add_argument('-g',dest='grade',help="电子凭证开户类别（开户模式选 b ，此参数才有效；默认 1）：\n    1: 核心企业\n    2: 一级供应商\n    3: 二级供应商\n    0: 以上三种类型",required=False,default='1',choices=['1','2','3','0'])
	parser.add_argument('-l',dest='loadCompany',help="使用本地存储的company对象（只有开户模式选择 b 此参数才有效）",required=False)
	parser.add_argument("-i",dest='legalName',help="使用指定人的身份信息（默认 *）",required=False,default='*')
	parser.add_argument('-e',dest='env',help="运行环境（默认 c）：\n    k: 开发(46)环境\n    c: 测试环境\n    p: 准生产环境\n    s: 生产环境",required=False,default='c',choices=['k','c','p','s'])
	parser.add_argument("-t",dest='type',help="注册类型（默认 2）：\n    1: 企业商户 - - 普通营业执照\n    2: 企业商户 - - 三证合一营业执照\n    3: 企业商户 - - 个体工商户营业执照\n    0: 注册以上三种类型",required=False,default='2',choices=['0','1','2','3'])
	parser.add_argument("-s",'--skip',help="开户时跳过绑定银行卡",action='store_true',required=False)
	parser.add_argument("-p",dest='precredit',help="是否自动预授信（默认 y）:\n    y: 是\n    n: 否",required=False,default='y',choices=['y','n'])
	parser.add_argument("-d",dest='daddy',help="选择授信方（默认 1）:\n    1: *\n    2: *\n    3: *",required=False,default='1')
	parser.add_argument("-c",dest='creditType',help="自动预授信类型（默认 1,可组合）:\n    1: *\n    2: *\n    3: *",required=False,default='1')
	parser.add_argument("-a",dest='all',help="是否做全套流程（默认 y）:\n    y: 是\n    n: 否（不做新资质审核、申请白条）",required=False,default='y',choices=['y','n'])
	parser.add_argument("-r",dest='registNum',help="使用指定的营业执照号(15位)",required=False)
	parser.add_argument("-u",dest='uscnum',help="使用指定的统一社会信用代码(18位)",required=False)
	parser.add_argument("-o",'--overfill',help="启用边界值模式",action='store_true',required=False)
	parser.add_argument('--auditSta',help="设置大总管审核不通过",action='store_true',required=False)
	parser.add_argument("--getRegUrl",help="获得注册链接并退出",action='store_true',required=False)
	parser.add_argument("--path",help=argparse.SUPPRESS,required=False,default='')
	parser.add_argument("--robot",help=argparse.SUPPRESS,action='store_true',required=False)
	args=parser.parse_args()
	return args

def main(**bro):
	try:
		global newpassword,paypassword,manHead,liaHead,browserFund
		if args.getRegUrl:
			longUrl=f"{domain}/bestpay/register?{encryptRegData(company.loginID,time.strftime('%Y%m%d%H%M%S',time.localtime()),platCode)}"
			shortUrl=getShortUrl(longUrl)
			if shortUrl:
				print(shortUrl)
			else:
				print(longUrl)
			print()
			sys.exit()
		if bro:
			browser=bro['bro']
			browser.get(f"{domain}/bestpay/register?{encryptRegData(company.loginID,time.strftime('%Y%m%d%H%M%S',time.localtime()),platCode,company.email)}")
		else:
			browser=openurl(f"{domain}/bestpay/register?{encryptRegData(company.loginID,time.strftime('%Y%m%d%H%M%S',time.localtime()),platCode,company.email)}")
		if args.mode=='b':
			print('-'*50,'注册:',gradeDic[args.grade],'-'*50)
			if args.robot:saveFileRobot(f"{'-'*50} 注册:{gradeDic[args.grade]} {'-'*50}")
		print(f"PC {envDict[args.env]} {time.strftime('%Y-%m-%d %X',time.localtime())}")
		if args.robot:saveFileRobot(f"PC {envDict[args.env]} {time.strftime('%Y-%m-%d %X',time.localtime())}")
		if saveToMysql:mysqlOpt(f"""INSERT INTO registerinfo (CreateDate,Environmental,Grade,loginID,RegistTypes) VALUES ("{time.strftime('%Y-%m-%d %X',time.localtime())}","PC {envDict[args.env]}","{gradeDic[args.grade]}","{company.loginID}","{regTypeDic[args.type]}")""")
		else:save_file(f'登录号： {company.loginID}\n{regTypeDic[args.type]}')
		print(f'登录号： {company.loginID}\n{regTypeDic[args.type]}')	
		if args.robot:saveFileRobot(f'登录号： {company.loginID}\n{regTypeDic[args.type]}')	
		# if saveToMysql:mysqlOpt(f"UPDATE registerinfo SET RegistTypes='{regTypeDic[args.type]}' WHERE LoginID='{company.loginID}'")
		# else:save_file(regTypeDic[args.type])
		# print(regTypeDic[args.type])
		if args.type=='1':browser=fillbankinfor(postpic0(beneficiaryInfo(legalinformation(Putong(browser)))),0)
		elif args.type=='2':browser=fillbankinfor(postpic1(beneficiaryInfo(legalinformation(Sanzhengheyi(browser)))),1)
		elif args.type=='3':browser=fillbankinfor(postpic2(beneficiaryInfo(legalinformation(Getigongshanghu(browser)))),2)
		elif args.type=='4':browser=fillbankinfor(postpic3(Geti(browser)),3)
		if saveToMysql:mysqlOpt(f"UPDATE registerinfo SET IsReview=0,IsPreCredit=0,IsApplicate=0 WHERE LoginID='{company.loginID}'")
		if args.noCheck:#小CEO批量模式
			# browser.quit()
			return '提交完成',browser
		if args.robot:saveFileRobot('已提交开户申请，下一步大总管审核……')
		# browser=bbossCheck(company.loginID,env=args.env,bro=browser)
		out('大总管审核中……\t\t\t')
		auditSta=3 if args.auditSta else 2
		bbCheckData=bbossCheck(company.loginID,env=args.env,op='1',auditSta=auditSta,path=args.path)
		if bbCheckData['auditSta']=='审核通过':
			out('大总管审核通过……\t\t\t')
		else:
			print('大总管',bbCheckData['auditSta'],bbCheckData['respCode'],bbCheckData['respMsg'])
			sys.exit()
		if saveToMysql:mysqlOpt(f"UPDATE registerinfo SET IsReview=1 WHERE LoginID='{company.loginID}'")
		if args.robot:saveFileRobot('大总管已审核，下一步获取初始密码……')
		if args.env in ['k','c']:
			# browser,om_cookie=loginOM(env=args.env,bro=browser)
			out('正在获取初始密码...')
			password=getOMpassword(company.loginID,company.legalTel,env=args.env,path=args.path)#获取初始密码
			out('正在获取商户号...')
			merchantNum=getMerchantNum(company.loginID,env=args.env,path=args.path)#获取商户号
			if saveToMysql:mysqlOpt(f"UPDATE registerinfo SET MerchantNum='{merchantNum}' WHERE LoginID='{company.loginID}'")
		elif args.env in ['p','s']:
			if args.robot:
				sendInput('请发送初始密码，格式：初始密码 xxxx')
				password=getInput('初始密码')
			else:
				password=input('输入初始密码后按回车，退出请按"q"：')
				if password=='q':
					browser.quit()
					sys.exit(0)
		if saveToMysql:mysqlOpt(f"UPDATE registerinfo SET InitialPWD='{password}' WHERE LoginID='{company.loginID}'")
		else:save_file(f'初始密码： {password}')
		print(f'初始密码： {password}\t\t\t\t')
		if args.robot:saveFileRobot(f'初始密码： {password}')
		if noPic or args.env=='k':#H5没有开发环境
		# if 1==1:
			browser=changePassword(company.loginID,password,newpassword,paypassword,env=args.env,bro=browser)
		else:#不禁止加载图片则可以去H5页面修改密码
			browser1=loginH5(username=company.loginID,passwd=password,env=args.env,menhuPath=args.path)#初始密码登陆
			browser1,newpassword=changePasswordH5(loginId=company.loginID,password=password,newpassword=newpassword,paypassword=paypassword,env=args.env,menhuPath=args.path,bro=browser1)
			browser1.quit()
		if saveToMysql:mysqlOpt(f"UPDATE registerinfo SET Password='{newpassword}',Payword='{paypassword}' WHERE LoginID='{company.loginID}'")
		else:save_file(f'登陆密码： {newpassword}\t支付密码： {paypassword}')
		print(f'登陆密码： {newpassword}\t支付密码： {paypassword}')
		if args.robot:saveFileRobot(f'登陆密码： {newpassword}\t支付密码： {paypassword}')
		result='密码修改完成！'
		creditTypeDic={'1':'佣金','2':'采购','3':'小CEO',}
		if args.mode=='a':
			if args.precredit=='y':
				creditHead=0
				for creditType in args.creditType:
					out(f'正在导入预授信，类型 {creditTypeDic[creditType]}...')
					save_excel(company.BLRN,company.companyName,creditType,args.daddy)
					creditResult,creditHead=uploadExcel(idlist=[company.BLRN],env=args.env,head=creditHead,menhuPath=args.path)
					addWhiteList(args.env,creditHead,company,args.daddy,args.robot)
					if not creditResult:#预授信
						if args.robot:
							print(f'导入预授信失败，类型 {creditTypeDic[creditType]}')
							sys.exit()
						else:
							s=input(f'导入预授信失败，类型 {creditTypeDic[creditType]} ,请手动完成后按回车，退出请按"q"：')
							if s=='q':
								browser.quit()
								sys.exit(0)
			else:
				if args.robot:
					sendInput('请手动导入预授信后发送 已完成预授信')
					getInput('已完成预授信')
				else:
					s=input('请手动完成预授信后按回车，退出请按"q"：')
					if s=='q':
						browser.quit()
						sys.exit(0)
			if saveToMysql:mysqlOpt(f"UPDATE registerinfo SET IsPreCredit=1 WHERE LoginID='{company.loginID}'")
			else:save_file('已完成预授信')
			print('已完成预授信\t\t')
			if args.robot:saveFileRobot('已完成预授信\t\t')
			if args.all=='y':
				JBC=True if args.daddy=='2' else False
				browser,result=AutoFill(company.loginID,newpassword,env=args.env,bankCode=company.bankCode,oFill=args.overfill,pLegalID=company.getPlegalID(company.legalID),JBC=JBC,robot=args.robot,path=args.path,bro=login(company.loginID,newpassword,env=args.env,menhuPath=args.path,bro=browser))
				if saveToMysql:mysqlOpt(f"UPDATE registerinfo SET IsApplicate=1 WHERE LoginID='{company.loginID}'")
				else:save_file(f"完成！ {time.strftime('%Y-%m-%d %X',time.localtime())}")
			browser.quit()
			return f'{result}\t\t\t\t'
		elif args.mode=='b':
			browser=login(username=company.loginID,passwd=newpassword,env=args.env,menhuPath=args.path,bro=browser)#改完密码后登陆
			browser,initialPWD=addOperator(opLoginID=company.opLoginID,env=args.env,newRole=1,bro=browser)#新增操作员
			if saveToMysql:mysqlOpt(f"""INSERT INTO registerinfo (CreateDate,Environmental,Grade,loginID,RegistTypes,InitialPWD) VALUES ("{time.strftime('%Y-%m-%d %X',time.localtime())}","PC {envDict[args.env]}","{gradeDic[args.grade]}","{company.opLoginID}","操作员","{initialPWD}")""")
			else:save_file(f'操作员登录号： {company.opLoginID}')
			print(f'操作员登录号： {company.opLoginID}\t初始密码：{initialPWD}')
			if args.robot:saveFileRobot(f'操作员登录号： {company.opLoginID}\t初始密码：{initialPWD}')
			if noPic or args.env=='k':#H5没有开发环境
				browser=changePassword(company.opLoginID,initialPWD,newpassword,paypassword,env=args.env,bro=browser)
			else:#不禁止加载图片则可以去H5页面修改密码
				# browser.quit()
				browser1=loginH5(username=company.opLoginID,passwd=initialPWD,env=args.env,menhuPath=args.path)#初始密码登陆
				browser1,newpassword=changePasswordH5(loginId=company.opLoginID,password=initialPWD,newpassword=newpassword,paypassword=paypassword,env=args.env,menhuPath=args.path,bro=browser1)
				browser1.quit()
			if saveToMysql:mysqlOpt(f"UPDATE registerinfo SET Password='{newpassword}',Payword='{paypassword}' WHERE LoginID='{company.opLoginID}'")
			else:save_file(f'操作员登陆密码： {newpassword}\t操作员支付密码： {paypassword}')
			print(f'操作员登陆密码： {newpassword}\t操作员支付密码： {paypassword}')
			if args.robot:saveFileRobot(f'操作员登陆密码： {newpassword}\t操作员支付密码： {paypassword}')
			#电子凭证授信
			company.cusType='CORE' if args.grade=='1' else 'CUSTOMER'
			browser.maximize_window()
			browser,result,manHead,liaHead=AutoFillElect(company=company,renewal='n',reFill='y',passwd=newpassword,bro=browser)
			print(f'{result}\t\t\t\t')
			if args.robot:saveFileRobot(result)
			# 翼融平台客户录入
			browser.find_element_by_id('login').click()#登出
			if browserFund:
				browserFund=entryCustInfo(company=company,jump=0,needLogin=0,bro=browserFund)#换资金方经办员录入客户信息
			else:
				browserFund=entryCustInfo(company=company,jump=1,needLogin=1)#换资金方经办员录入客户信息
			print('翼融平台客户录入成功！\t\t\t\t\t\n')
			if args.robot:saveFileRobot('翼融平台客户录入成功！')
			return browser,browserFund
	except Exception as err:
		reason=getErrReason(browser)
		if not reason:reason=err
		print('开户失败！原因：',reason)
		sys.exit()

if __name__ == '__main__':
	envDict={'k':'开发','c':'测试','s':'生产','p':'准生产'}
	platCodeDic={'k':'*','c':'*','s':'*','p':'*'}
	gradeDic={'1':'核心企业','2':'一级供应商','3':'二级供应商',None:'-'}
	regTypeDic={'1':'企业商户--普通执照','2':'企业商户--三证合一','3':'企业商户--个体工商'}
	industryList=['通讯类','零售连锁类','证券公司','物流运输','商旅类','保险公司','电商类','外贸','教育医疗','其他']
	args=getParserPC()
	if args.env=='c':
		if not args.getRegUrl:
			build=checkBuild()
			if build:
				if '失败' in build:print(build)
				else:print(f'检测到在发版: {build} 请稍后再试！')
				if args.robot:saveFileRobot(f'检测到正在发版: {build} 请稍后再试！')
				sys.exit()
	elif args.env=='s':
		if not args.getRegUrl:
			head={
				'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/75.0.3770.100 Safari/537.36',
				'Content-Type':'application/x-www-form-urlencoded; charset=UTF-8',
				'Cookie':open(os.path.join(args.path,f'cookie\\BB_s'),'r',encoding='UTF-8').read(),
				'Connection':'close',
			}
			if not BBalive(head,'s'):
				if args.robot:
					sendInput('检测到生产大总管cookie已失效，请更新后重试。查看更新方法请发送 更新cookie')
					sys.exit()
				else:
					print('检测到生产大总管cookie已失效，请更新后重试.')
			head['Cookie']=open(os.path.join(args.path,f'cookie\\MM_s'),'r',encoding='UTF-8').read()
			if not isMMalive(head,'s'):
				if args.robot:
					sendInput('检测到生产资金后台cookie已失效，请更新后重试。查看更新方法请发送 更新cookie')
					sys.exit()
				else:
					print('检测到生产资金后台cookie已失效，请更新后重试.')
	if args.robot:
		try:os.remove(RFP)
		except FileNotFoundError:pass
	args.newpassword=newpassword;args.paypassword=paypassword
	args.noCheck=0
	platCode=platCodeDic[args.env];domain=domainDic[args.env]
	flag=1 if args.type=='0' or args.grade=='0' else 0
	if args.env in ['k','c','p','s'] and args.type in ['1','2','3','0']:
		browserFund=0
		start=time.perf_counter()
		if flag:#批量开
			custCreditIdList=[];companyList=[]
			browser=0
			if not args.loadCompany:
				for i in range(1,4):
					args.type=str(i);args.grade=str(i)
					company=GenCompanyInfo('pc',args.env,args.legalName)
					if args.grade=='1':CoreCreditId=company.BLRN
					else:custCreditIdList.append(company.BLRN)
					if args.type=='1':company.BLRN=''.join([str(randint(0,9)) for i in range(15)])#营业执照注册号,15位数字,随机生成
					if args.mode=='a':
						company.legalName,company.legalID,company.bankCode,company.legalTel=getLegalInfo(args.env,args.legalName)#legalName来自config.py
						company.bankAccName=company.legalName
						result=main()
						print(result)
						if args.robot:saveFileRobot(result)
					elif args.mode=='b':
						if browser:browser,browserFund=main(bro=browser)
						else:browser,browserFund=main()
						companyList.append(company)
						company.save()
				browser.quit()
			if args.mode=='b':
				if args.loadCompany:
					for loginID in args.loadCompany.split():
						company=loadCompany(loginID)
						companyList.append(company)
						if company.cusType=='CORE':CoreCreditId=company.BLRN
						else:custCreditIdList.append(company.BLRN)
					browserFund,productName=addProduct(jump=1,needLogin=1,env=args.env)#资金方经办新增产品
				else:
					browserFund,productName=addProduct(jump=0,needLogin=0,env=args.env,bro=browserFund)#资金方经办新增产品
				browserFund=addProductChain(CoreCreditId,custCreditIdList,jump=0,needLogin=0,env=args.env,bro=browserFund)#资金方经办新增产业链
				browserFund=addProject(CoreCreditId,custCreditIdList,productName,jump=0,needLogin=0,env=args.env,bro=browserFund)#资金方经办新增方案
				browserFund=addCredit(CoreCreditId,custCreditIdList,jump=0,needLogin=0,env=args.env,bro=browserFund)#资金方经办录入额度
				# head=''
				for company in companyList:
					manHead=addUser(company,manHead)#业务经理新增用户、角色
					print(company.companyName,'新增角色、用户成功！')
					if args.robot:saveFileRobot(f'{company.companyName} 新增角色、用户成功！')
					if company.cusType=='CORE':
						entryCoreCompany(company,manHead)#业务经理录入核心企业
						print(company.companyName,'录入核心企业成功！')
						if args.robot:saveFileRobot(f'{company.companyName} 录入核心企业成功！')
		else:#单开
			company=GenCompanyInfo('pc',args.env,args.legalName)
			if args.type=='1':company.BLRN=''.join([str(randint(0,9)) for i in range(15)])#营业执照注册号,15位数字,随机生成
			if args.registNum:#指定普通执照
				if args.type=='1':
					if len(args.registNum)==15:
						company.BLRN=args.registNum
					else:
						print('错误: 输入的营业执照号长度不是15位！')
						if args.robot:saveFileRobot('错误: 输入的营业执照号长度不是15位！')
						sys.exit()
				elif args.type=='3':
					company.BLRN=args.registNum
				else:
					print(f'错误: 注册类型为 {regTypeDic[args.type]} 时不能指定营业执照号！')
					if args.robot:saveFileRobot(f'错误: 注册类型为 {regTypeDic[args.type]} 时不能指定营业执照号！')
					sys.exit()
			elif args.uscnum:#指定统一社会信用代码
				if args.type=='2':
					if len(args.uscnum)==18:
						company.BLRN=args.uscnum
					else:
						print('错误: 输入的统一社会信用代码长度不是18位！')
						if args.robot:saveFileRobot('错误: 输入的统一社会信用代码长度不是18位！')
						sys.exit()
				elif args.type=='3':
					company.BLRN=args.registNum
				else:
					print(f'错误: 注册类型为 {regTypeDic[args.type]} 时不能指定统一社会信用代码！')
					if args.robot:saveFileRobot(f'错误: 注册类型为 {regTypeDic[args.type]} 时不能指定统一社会信用代码！')
					sys.exit()
			if args.mode=='a':
				company.legalName,company.legalID,company.bankCode,company.legalTel=getLegalInfo(args.env,args.legalName)#legalName来自config.py
				company.bankAccName=company.legalName
				result=main()
				print(result)
				if args.robot:saveFileRobot(result)
			elif args.mode=='b':
				main()
				company.save()
				browserFund.quit()
			elif args.mode=='c':
				from openpyxl import load_workbook
				timestr=time.strftime('%Y%m%d%H%M%S',time.localtime())
				excel=load_workbook('info/infomations.xlsx')
				sheet=excel.active
				regTypeDicDesc={'普通营业执照':'1','三证合一营业执照':'2','个体工商户营业执照':'3'}
				chrome_option=webdriver.ChromeOptions()
				chrome_option.add_argument('disable-infobars')
				browser=webdriver.Chrome(options=chrome_option)
				browser.maximize_window()
				for row in sheet.iter_rows(min_row=2):
					saveToMysql=0
					# args.skip='y'
					args.noCheck=1
					args.type=regTypeDicDesc[getCellStr(row[0])]
					company.companyName=getCellStr(row[1])
					company.BLRN=getCellStr(row[2])
					company.province,company.city,company.district=getCellStr(row[3]).split()
					company.province,company.city,company.district=getArea(company.province,company.city,company.district)
					company.officeAddr=getCellStr(row[4])
					company.industry=getCellStr(row[5])
					company.regTime,company.endTime=getCellStr(row[6]).split('至')
					company.permitNum=getCellStr(row[7])
					company.scope=getCellStr(row[8])
					company.regCapital=getCellStr(row[9])
					company.legalName=getCellStr(row[12])
					company.legalID=getCellStr(row[14])
					company.LICST,company.LICET=getCellStr(row[15]).split('至')
					company.legalTel=getCellStr(row[16])
					# company.email=getCellStr(row[17])
					company.email=f'{company.legalTel}@189.cn'
					company.beneName=getCellStr(row[24])
					company.beneID=getCellStr(row[26])
					try:company.BICST,company.BICET=getCellStr(row[27]).split('至')
					except (AttributeError,ValueError):pass
					company.beneAddr=getCellStr(row[28])
					company.bizLicensepath=os.path.join(os.getcwd(),f'info/{company.BLRN}/1.jpg')#营业执照
					if not os.path.exists(company.bizLicensepath):company.bizLicensepath=''
					company.legalApath=os.path.join(os.getcwd(),f'info/{company.BLRN}/2.jpg')#身份证正面
					if not os.path.exists(company.legalApath):company.legalApath=''
					company.legalBpath=os.path.join(os.getcwd(),f'info/{company.BLRN}/3.jpg')#身份证反面
					if not os.path.exists(company.legalBpath):company.legalBpath=''
					company.handIDpath=os.path.join(os.getcwd(),f'info/{company.BLRN}/4.jpg')#手持身份证
					if not os.path.exists(company.handIDpath):company.handIDpath=''
					company.openAcctpath=os.path.join(os.getcwd(),f'info/{company.BLRN}/5.jpg')#银行开户许可证
					if not os.path.exists(company.openAcctpath):company.openAcctpath=''
					# company.orgCodepath=os.path.join(os.getcwd(),f'info/{company.BLRN}/组织机构代码证.jpg')
					# company.taxRegpath=os.path.join(os.getcwd(),f'info/{company.BLRN}/税务登记证.jpg')
					# company.bankCardpath=os.path.join(os.getcwd(),f'info/{company.BLRN}/对公银行开户回执单.jpg')
					company.loginID=f'AAA{company.legalTel}'
					platCode='*'
					try:
						result,browser=main(bro=browser)
						print(company.companyName,result,'\n')
					except Exception as err:
						reason=getErrReason(browser)
						print('错误：',company.BLRN,'原因：',reason)
						with open(f'info/ErrorID{timestr}.txt','a+',encoding='UTF-8') as f:f.write(f'错误: {company.BLRN} 原因: {reason}\n')
						if not reason:os.system('pause')
						print()
				browser.quit()
		if os.path.exists('geckodriver.log'):os.remove('geckodriver.log')
		t=time.strftime('%M{m}%S{s}',time.gmtime(time.perf_counter()-start)).format(m='分',s='秒')
		print(f"完成！ {time.strftime('%Y-%m-%d %X',time.localtime())}\n用时：{t}")
		if args.robot:saveFileRobot(f"完成！ {time.strftime('%Y-%m-%d %X',time.localtime())}\n用时：{t}")
	else:
		print('输入有误！键入 "regPc.py -h" 查看帮助！')
		if args.robot:saveFileRobot('输入有误！键入 "regPc.py -h" 查看帮助！')
		sys.exit()